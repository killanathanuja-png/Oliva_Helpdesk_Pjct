from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.models import ChildCategory, Subcategory, Category, StatusEnum
from app.schemas.schemas import ChildCategoryCreate, ChildCategoryUpdate, ChildCategoryResponse
import io

router = APIRouter(prefix="/api/child-categories", tags=["ChildCategories"])


def _next_code(db: Session) -> str:
    last = db.query(ChildCategory).order_by(ChildCategory.id.desc()).first()
    num = (last.id + 1) if last else 1
    return f"CC{num:03d}"


@router.get("/", response_model=list[ChildCategoryResponse])
def list_child_categories(db: Session = Depends(get_db)):
    items = db.query(ChildCategory).order_by(ChildCategory.id).all()
    return [
        ChildCategoryResponse(
            id=c.id, code=c.code, name=c.name,
            subcategory=c.subcategory_rel.name if c.subcategory_rel else None,
            category=c.category_rel.name if c.category_rel else None,
            module=c.module,
            status=c.status.value if c.status else "Active",
            created_at=c.created_at,
        )
        for c in items
    ]


@router.post("/", response_model=ChildCategoryResponse, status_code=201)
def create_child_category(req: ChildCategoryCreate, db: Session = Depends(get_db)):
    sub = db.query(Subcategory).filter(Subcategory.name == req.subcategory).first() if req.subcategory else None
    cat = db.query(Category).filter(Category.name == req.category).first() if req.category else None

    code = req.code.strip() if req.code and req.code.strip() else _next_code(db)
    child = ChildCategory(
        code=code, name=req.name, module=req.module,
        subcategory_id=sub.id if sub else None,
        category_id=cat.id if cat else None,
        status=StatusEnum.Active,
    )
    db.add(child)
    db.commit()
    db.refresh(child)
    return ChildCategoryResponse(
        id=child.id, code=child.code, name=child.name,
        subcategory=req.subcategory, category=req.category,
        module=child.module,
        status="Active", created_at=child.created_at,
    )


@router.put("/{item_id}", response_model=ChildCategoryResponse)
def update_child_category(item_id: int, req: ChildCategoryUpdate, db: Session = Depends(get_db)):
    child = db.query(ChildCategory).filter(ChildCategory.id == item_id).first()
    if not child:
        raise HTTPException(status_code=404, detail="Child Category not found")
    if req.name is not None:
        child.name = req.name
    if req.module is not None:
        child.module = req.module
    if req.subcategory is not None:
        sub = db.query(Subcategory).filter(Subcategory.name == req.subcategory).first()
        child.subcategory_id = sub.id if sub else None
    if req.category is not None:
        cat = db.query(Category).filter(Category.name == req.category).first()
        child.category_id = cat.id if cat else None
    if req.status is not None:
        child.status = StatusEnum(req.status)
    db.commit()
    db.refresh(child)
    return ChildCategoryResponse(
        id=child.id, code=child.code, name=child.name,
        subcategory=child.subcategory_rel.name if child.subcategory_rel else None,
        category=child.category_rel.name if child.category_rel else None,
        module=child.module,
        status=child.status.value if child.status else "Active",
        created_at=child.created_at,
    )


@router.patch("/{item_id}/status")
def update_child_category_status(item_id: int, status: str, db: Session = Depends(get_db)):
    child = db.query(ChildCategory).filter(ChildCategory.id == item_id).first()
    if not child:
        raise HTTPException(status_code=404, detail="Child Category not found")
    child.status = StatusEnum(status)
    db.commit()
    db.refresh(child)
    return ChildCategoryResponse(
        id=child.id, code=child.code, name=child.name,
        subcategory=child.subcategory_rel.name if child.subcategory_rel else None,
        category=child.category_rel.name if child.category_rel else None,
        module=child.module,
        status=child.status.value if child.status else "Active",
        created_at=child.created_at,
    )


@router.delete("/{item_id}", status_code=204)
def delete_child_category(item_id: int, db: Session = Depends(get_db)):
    child = db.query(ChildCategory).filter(ChildCategory.id == item_id).first()
    if not child:
        raise HTTPException(status_code=404, detail="Child Category not found")
    db.delete(child)
    db.commit()


@router.post("/upload-excel")
def upload_child_categories_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Bulk upload child categories from Excel."""
    import openpyxl

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file. Please upload a valid .xlsx file.")

    ws = wb.active
    if not ws or ws.max_row < 2:
        raise HTTPException(status_code=400, detail="Excel file is empty or has no data rows.")

    # Read headers
    raw_headers = [str(ws.cell(1, c).value or "").strip().lower().replace("\n", "").replace(" ", "_") for c in range(1, ws.max_column + 1)]

    # Map headers
    header_map = {}
    aliases = {
        "code": ["child_category_code", "code", "child_code", "childcategorycode"],
        "name": ["child_category_name", "name", "child_name", "childcategoryname"],
        "subcategory": ["sub_category_name", "sub_category", "subcategory", "subcategoryname", "sub_category_code"],
        "category": ["main_category", "category", "main_category_name", "maincategory"],
        "module": ["module"],
        "status": ["status"],
    }
    for field, possible in aliases.items():
        for i, h in enumerate(raw_headers):
            if h in possible:
                header_map[field] = i
                break

    if "name" not in header_map and "code" not in header_map:
        raise HTTPException(status_code=400, detail="Excel must have at least 'Child Category Name' or 'Child Category Code' column.")

    # Build lookups
    sub_map = {s.name.lower(): s.id for s in db.query(Subcategory).all()}
    cat_map = {c.name.lower(): c.id for c in db.query(Category).all()}
    existing_codes = {c.code: c for c in db.query(ChildCategory).all()}

    last = db.query(ChildCategory).order_by(ChildCategory.id.desc()).first()
    code_counter = (last.id + 1) if last else 1

    added = 0
    updated = 0
    skipped = 0

    def cell_val(row_idx: int, field: str) -> str:
        col_idx = header_map.get(field)
        if col_idx is None:
            return ""
        val = ws.cell(row_idx, col_idx + 1).value
        if val is None:
            return ""
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        return str(val).strip()

    for row_num in range(2, ws.max_row + 1):
        name = cell_val(row_num, "name")
        code = cell_val(row_num, "code")

        if not name and not code:
            continue

        if not name:
            name = code

        subcategory_name = cell_val(row_num, "subcategory")
        category_name = cell_val(row_num, "category")
        module = cell_val(row_num, "module")
        status_val = cell_val(row_num, "status") or "Active"

        sub_id = sub_map.get(subcategory_name.lower()) if subcategory_name else None
        cat_id = cat_map.get(category_name.lower()) if category_name else None

        if code and code in existing_codes:
            # Update existing
            child = existing_codes[code]
            child.name = name
            if sub_id is not None:
                child.subcategory_id = sub_id
            if cat_id is not None:
                child.category_id = cat_id
            if module:
                child.module = module
            if status_val in ("Active", "Inactive"):
                child.status = StatusEnum(status_val)
            updated += 1
        else:
            # Create new
            if not code:
                code = f"CC{code_counter:03d}"
                code_counter += 1
            child = ChildCategory(
                code=code, name=name, module=module or None,
                subcategory_id=sub_id, category_id=cat_id,
                status=StatusEnum(status_val) if status_val in ("Active", "Inactive") else StatusEnum.Active,
            )
            db.add(child)
            existing_codes[code] = child
            added += 1

    db.commit()

    # Return updated list
    items = db.query(ChildCategory).order_by(ChildCategory.id).all()
    item_list = [
        ChildCategoryResponse(
            id=c.id, code=c.code, name=c.name,
            subcategory=c.subcategory_rel.name if c.subcategory_rel else None,
            category=c.category_rel.name if c.category_rel else None,
            module=c.module,
            status=c.status.value if c.status else "Active",
            created_at=c.created_at,
        )
        for c in items
    ]

    return {
        "message": f"Added {added}, Updated {updated}, Skipped {skipped}",
        "added": added, "updated": updated, "skipped": skipped,
        "items": item_list,
    }
