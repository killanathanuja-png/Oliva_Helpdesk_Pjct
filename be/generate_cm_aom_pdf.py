"""Generate a PDF and Excel sheet grouped by AOM (Area Operations Manager),
listing each AOM with their email and the centers they manage along with
each center's Center Manager (CM) name and email.

Outputs (in the `be/` directory):
  - cm_aom_allocation.pdf
  - cm_aom_allocation.xlsx
"""
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)

from app.database import SessionLocal
from app.models.models import AOMCenterMapping


PDF_FILE = "cm_aom_allocation.pdf"
XLSX_FILE = "cm_aom_allocation.xlsx"
OUTPUT_FILE = PDF_FILE  # backward-compat alias


def fetch_grouped():
    db = SessionLocal()
    try:
        rows = (
            db.query(AOMCenterMapping)
            .order_by(AOMCenterMapping.aom_name, AOMCenterMapping.center_name)
            .all()
        )
    finally:
        db.close()

    grouped = defaultdict(lambda: {"email": "", "centers": []})
    for r in rows:
        key = r.aom_name or ""
        grouped[key]["email"] = r.aom_email or grouped[key]["email"]
        grouped[key]["centers"].append({
            "center": r.center_name or "",
            "cm_name": r.cm_name or "",
            "cm_email": r.cm_email or "",
        })
    return grouped


def build_pdf(grouped):
    doc = SimpleDocTemplate(
        PDF_FILE,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="AOM Allocation with Centers and CMs",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], fontSize=16, spaceAfter=6
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontSize=9, leading=11
    )

    total_centers = sum(len(v["centers"]) for v in grouped.values())
    elements = [
        Paragraph("AOM Allocation — Centers &amp; Center Managers", title_style),
        Paragraph(
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} — "
            f"{len(grouped)} AOM(s), {total_centers} center mapping(s)",
            sub_style,
        ),
        Spacer(1, 8),
    ]

    header = ["AOM Name", "AOM Email", "Allocated Center", "CM Name", "CM Email"]
    data = [header]

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d9488")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, 0), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]

    row_idx = 1
    band = 0
    for aom_name in sorted(grouped.keys()):
        group = grouped[aom_name]
        centers = group["centers"]
        if not centers:
            continue
        first_row = row_idx
        for i, c in enumerate(centers):
            if i == 0:
                data.append([
                    Paragraph(aom_name, cell_style),
                    Paragraph(group["email"], cell_style),
                    Paragraph(c["center"], cell_style),
                    Paragraph(c["cm_name"], cell_style),
                    Paragraph(c["cm_email"], cell_style),
                ])
            else:
                data.append([
                    "", "",
                    Paragraph(c["center"], cell_style),
                    Paragraph(c["cm_name"], cell_style),
                    Paragraph(c["cm_email"], cell_style),
                ])
            row_idx += 1
        last_row = row_idx - 1
        # Span AOM name & email across the AOM's centers
        if last_row > first_row:
            style_cmds.append(("SPAN", (0, first_row), (0, last_row)))
            style_cmds.append(("SPAN", (1, first_row), (1, last_row)))
        # Alternating band per AOM block for readability
        bg = colors.white if band % 2 == 0 else colors.HexColor("#f1f5f9")
        style_cmds.append(("BACKGROUND", (0, first_row), (-1, last_row), bg))
        band += 1

    col_widths = [40 * mm, 60 * mm, 60 * mm, 40 * mm, 60 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)


def build_xlsx(grouped):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AOM Allocation"

    header = ["AOM Name", "AOM Email", "Allocated Center", "CM Name", "CM Email"]
    ws.append(header)

    header_fill = PatternFill("solid", fgColor="0D9488")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    band_fill = PatternFill("solid", fgColor="F1F5F9")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    row = 2
    band = 0
    merge_ranges = []
    for aom_name in sorted(grouped.keys()):
        group = grouped[aom_name]
        centers = group["centers"]
        if not centers:
            continue
        first_row = row
        for i, c in enumerate(centers):
            ws.cell(row=row, column=1, value=aom_name if i == 0 else None)
            ws.cell(row=row, column=2, value=group["email"] if i == 0 else None)
            ws.cell(row=row, column=3, value=c["center"])
            ws.cell(row=row, column=4, value=c["cm_name"])
            ws.cell(row=row, column=5, value=c["cm_email"])
            for col_idx in range(1, 6):
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = center_align
                cell.border = border
                if band % 2 == 1:
                    cell.fill = band_fill
            row += 1
        last_row = row - 1
        if last_row > first_row:
            merge_ranges.append((first_row, last_row))
        band += 1

    for first, last in merge_ranges:
        ws.merge_cells(start_row=first, start_column=1, end_row=last, end_column=1)
        ws.merge_cells(start_row=first, start_column=2, end_row=last, end_column=2)

    widths = [22, 38, 28, 24, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    wb.save(XLSX_FILE)


def main():
    grouped = fetch_grouped()
    if not grouped:
        print("No AOM/CM mappings found in aom_center_mappings table.")
        return
    build_pdf(grouped)
    build_xlsx(grouped)
    total = sum(len(v["centers"]) for v in grouped.values())
    print(f"Wrote {PDF_FILE} and {XLSX_FILE} — {len(grouped)} AOM(s), {total} center mapping(s).")


if __name__ == "__main__":
    main()
