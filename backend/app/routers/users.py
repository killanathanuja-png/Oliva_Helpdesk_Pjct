from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from app.database import get_db
from app.models.models import User, Department, Center, StatusEnum, user_centers
from app.schemas.schemas import UserCreate, UserUpdate, UserResponse
from app.auth import hash_password
import io

router = APIRouter(prefix="/api/users", tags=["Users"])


def _next_code(db: Session) -> str:
    last = db.query(User).order_by(User.id.desc()).first()
    num = (last.id + 1) if last else 1
    return f"U{num:03d}"


def _make_avatar(name: str) -> str:
    parts = name.strip().split()
    if len(parts) >= 2:
        return (parts[0][0] + parts[-1][0]).upper()
    return name[:2].upper()


def _to_response(u, dept_name=None, center_name=None) -> UserResponse:
    return UserResponse(
        id=u.id, code=u.code, employee_id=u.employee_id, name=u.name, email=u.email,
        role=u.role or None, map_level_access=u.map_level_access,
        designation=u.designation, entity=u.entity, vertical=u.vertical, costcenter=u.costcenter,
        department=dept_name if dept_name is not None else (u.department_rel.name if u.department_rel else None),
        center=center_name if center_name is not None else (u.center_rel.name if u.center_rel else None),
        gender=u.gender, mobile=u.mobile, reporting_to=u.reporting_to,
        grade=u.grade, employee_type=u.employee_type,
        city=u.city, employee_dob=u.employee_dob, employee_doj=u.employee_doj,
        lwd=u.lwd, effective_date=u.effective_date, remarks=u.remarks,
        avatar=u.avatar, status=u.status.value if u.status else None,
        last_login=u.last_login, created_at=u.created_at,
    )


@router.get("/")
def list_users(db: Session = Depends(get_db)):
    users = db.query(User).options(joinedload(User.department_rel), joinedload(User.center_rel)).order_by(User.id).all()
    result = []
    for u in users:
        try:
            result.append(_to_response(u))
        except Exception as e:
            import logging
            logging.error(f"Error serializing user {u.code} {u.name}: {e}")
    return result


@router.post("/", response_model=UserResponse, status_code=201)
def create_user(req: UserCreate, db: Session = Depends(get_db)):
    dept = db.query(Department).filter(Department.name == req.department).first() if req.department else None
    center = db.query(Center).filter(Center.name == req.center).first() if req.center else None

    emp_id = getattr(req, 'employee_id', None) or None
    code = emp_id if emp_id else _next_code(db)
    user = User(
        code=code, employee_id=emp_id,
        name=req.name, email=req.email,
        hashed_password=hash_password(req.password),
        role=req.role or "Employee", map_level_access=req.map_level_access,
        designation=req.designation, entity=req.entity, vertical=req.vertical, costcenter=req.costcenter,
        gender=req.gender, mobile=req.mobile, reporting_to=req.reporting_to,
        grade=req.grade, employee_type=req.employee_type,
        city=req.city, employee_dob=req.employee_dob, employee_doj=req.employee_doj,
        lwd=req.lwd, effective_date=req.effective_date, remarks=req.remarks,
        department_id=dept.id if dept else None,
        center_id=center.id if center else None,
        avatar=_make_avatar(req.name),
        status=StatusEnum.Active,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return _to_response(user, dept_name=req.department, center_name=req.center)


@router.get("/{user_id}", response_model=UserResponse)
def get_user(user_id: int, db: Session = Depends(get_db)):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    return _to_response(u)


@router.put("/{user_id}", response_model=UserResponse)
def update_user(user_id: int, req: UserUpdate, db: Session = Depends(get_db)):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    if req.name is not None:
        u.name = req.name
        u.avatar = _make_avatar(req.name)
    if req.email is not None:
        u.email = req.email
    if req.role is not None:
        u.role = req.role
    if req.map_level_access is not None:
        u.map_level_access = req.map_level_access
    if req.designation is not None:
        u.designation = req.designation
    if req.entity is not None:
        u.entity = req.entity
    if req.vertical is not None:
        u.vertical = req.vertical
    if req.costcenter is not None:
        u.costcenter = req.costcenter
    if req.department is not None:
        dept = db.query(Department).filter(Department.name == req.department).first()
        u.department_id = dept.id if dept else None
    if req.center is not None:
        center = db.query(Center).filter(Center.name == req.center).first()
        u.center_id = center.id if center else None
    if req.gender is not None:
        u.gender = req.gender
    if req.mobile is not None:
        u.mobile = req.mobile
    if req.reporting_to is not None:
        u.reporting_to = req.reporting_to
    if req.grade is not None:
        u.grade = req.grade
    if req.employee_type is not None:
        u.employee_type = req.employee_type
    if req.city is not None:
        u.city = req.city
    if req.employee_dob is not None:
        u.employee_dob = req.employee_dob
    if req.employee_doj is not None:
        u.employee_doj = req.employee_doj
    if req.lwd is not None:
        u.lwd = req.lwd
    if req.effective_date is not None:
        u.effective_date = req.effective_date
    if req.remarks is not None:
        u.remarks = req.remarks
    if req.status is not None:
        u.status = StatusEnum(req.status)
    db.commit()
    db.refresh(u)
    return _to_response(u)


@router.patch("/{user_id}/status", response_model=UserResponse)
def update_user_status(user_id: int, status: str, db: Session = Depends(get_db)):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    u.status = StatusEnum(status)
    db.commit()
    db.refresh(u)
    return _to_response(u)


@router.delete("/{user_id}", status_code=204)
def delete_user(user_id: int, db: Session = Depends(get_db)):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(u)
    db.commit()


@router.post("/upload-excel")
def upload_users_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Bulk upload users from an Excel file (.xlsx or .xls)."""
    import openpyxl

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file. Please upload a valid .xlsx file.")

    ws = wb.active
    if not ws or ws.max_row < 2:
        raise HTTPException(status_code=400, detail="Excel file is empty or has no data rows.")

    # Read headers (row 1)
    raw_headers = [str(ws.cell(1, c).value or "").strip().lower().replace("\n", "").replace(" ", "_") for c in range(1, ws.max_column + 1)]

    # Map header names to column indices
    header_map = {}
    aliases = {
        "user_id": ["user_id", "userid", "employee_id", "emp_id"],
        "user_name": ["user_name", "username", "name", "full_name"],
        "role_name": ["role_name", "role", "rolename"],
        "email_id": ["email_id", "email", "emailid", "email_address"],
        "map_level_access": ["map_level_access", "maplevelaccess", "access"],
        "gender": ["gender", "sex"],
        "designation": ["designation", "title"],
        "entity": ["entity"],
        "vertical": ["vertical"],
        "costcenter": ["costcenter", "cost_center"],
        "city": ["city"],
        "location": ["location", "center", "office"],
        "department": ["department", "dept"],
        "mobile_no": ["mobile_no", "mobile", "phone", "mobileno", "mobile_number"],
        "reporting_to": ["reporting_to", "reportingto", "reports_to", "manager"],
        "grade": ["grade"],
        "employee_type": ["employee_type", "employeetype", "emp_type"],
        "employee_dob": ["employee_dob", "employeedob", "dob", "date_of_birth"],
        "employee_doj": ["employee_doj", "employeedoj", "doj", "date_of_joining"],
        "lwd": ["lwd", "last_working_day"],
        "effective_date": ["effective_date", "effectivedate"],
        "remarks": ["remarks", "remark", "notes", "comment"],
        "status": ["status"],
        "password": ["password"],
    }

    for field, possible_names in aliases.items():
        for i, h in enumerate(raw_headers):
            if h in possible_names:
                header_map[field] = i
                break

    if "user_name" not in header_map and "email_id" not in header_map:
        raise HTTPException(status_code=400, detail="Excel must have at least 'User Name' or 'Email ID' column.")

    # Build lookup maps
    dept_map = {d.name.lower(): d.id for d in db.query(Department).all()}
    center_map = {c.name.lower(): c.id for c in db.query(Center).all()}
    existing_emails = {u.email: u for u in db.query(User).all()}
    last_user = db.query(User).order_by(User.id.desc()).first()
    code_counter = (last_user.id + 1) if last_user else 1

    added = 0
    updated = 0
    skipped = 0
    errors = []

    def cell_val(row_idx: int, field: str) -> str:
        col_idx = header_map.get(field)
        if col_idx is None:
            return ""
        val = ws.cell(row_idx, col_idx + 1).value
        if val is None:
            return ""
        # Handle datetime objects
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        return str(val).strip()

    for row_num in range(2, ws.max_row + 1):
        name = cell_val(row_num, "user_name")
        email = cell_val(row_num, "email_id").lower()

        # Skip completely empty rows
        if not name and not email:
            continue

        if not email:
            skipped += 1
            errors.append(f"Row {row_num}: No email, skipped")
            continue

        role = cell_val(row_num, "role_name") or "Employee"
        employee_id = cell_val(row_num, "user_id")
        password = cell_val(row_num, "password") or "oliva@123"
        dept_name = cell_val(row_num, "department")
        location = cell_val(row_num, "location")
        dept_id = dept_map.get(dept_name.lower()) if dept_name else None
        center_id = center_map.get(location.lower()) if location else None

        parts = name.split() if name else [email.split("@")[0]]
        avatar = (parts[0][0] + parts[-1][0]).upper() if len(parts) >= 2 else (name or email)[:2].upper()

        user_data = {
            "name": name or email.split("@")[0],
            "role": role,
            "employee_id": employee_id or None,
            "map_level_access": cell_val(row_num, "map_level_access") or None,
            "designation": cell_val(row_num, "designation") or None,
            "entity": cell_val(row_num, "entity") or None,
            "vertical": cell_val(row_num, "vertical") or None,
            "costcenter": cell_val(row_num, "costcenter") or None,
            "gender": cell_val(row_num, "gender") or None,
            "mobile": cell_val(row_num, "mobile_no") or None,
            "reporting_to": cell_val(row_num, "reporting_to") or None,
            "grade": cell_val(row_num, "grade") or None,
            "employee_type": cell_val(row_num, "employee_type") or None,
            "city": cell_val(row_num, "city") or None,
            "employee_dob": cell_val(row_num, "employee_dob") or None,
            "employee_doj": cell_val(row_num, "employee_doj") or None,
            "lwd": cell_val(row_num, "lwd") or None,
            "effective_date": cell_val(row_num, "effective_date") or None,
            "remarks": cell_val(row_num, "remarks") or None,
        }
        status_val = cell_val(row_num, "status")

        if email in existing_emails:
            # Update existing user
            u = existing_emails[email]
            for k, v in user_data.items():
                if v is not None:
                    setattr(u, k, v)
            if dept_id:
                u.department_id = dept_id
            if center_id:
                u.center_id = center_id
            u.avatar = avatar
            if status_val:
                try:
                    u.status = StatusEnum(status_val)
                except ValueError:
                    pass
            updated += 1
        else:
            # Create new user
            user_code = employee_id if employee_id else f"U{code_counter:03d}"
            code_counter += 1
            new_user = User(
                code=user_code, email=email,
                hashed_password=hash_password(password),
                department_id=dept_id, center_id=center_id,
                avatar=avatar, status=StatusEnum(status_val) if status_val and status_val in ("Active", "Inactive") else StatusEnum.Active,
                **user_data,
            )
            db.add(new_user)
            existing_emails[email] = new_user
            added += 1

    db.commit()

    # Return updated user list
    users = db.query(User).options(joinedload(User.department_rel), joinedload(User.center_rel)).order_by(User.id).all()
    user_list = []
    for u in users:
        try:
            user_list.append(_to_response(u))
        except Exception:
            pass

    return {
        "message": f"Added {added}, Updated {updated}, Skipped {skipped}",
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:10],
        "users": user_list,
    }


# --- Managed Centers (AOM multi-location) ---

@router.get("/{user_id}/managed-centers")
def get_managed_centers(user_id: int, db: Session = Depends(get_db)):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    return [{"id": c.id, "name": c.name} for c in u.managed_centers]


@router.put("/{user_id}/managed-centers")
def set_managed_centers(user_id: int, center_ids: list[int], db: Session = Depends(get_db)):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    centers = db.query(Center).filter(Center.id.in_(center_ids)).all()
    u.managed_centers = centers
    db.commit()
    return [{"id": c.id, "name": c.name} for c in u.managed_centers]
